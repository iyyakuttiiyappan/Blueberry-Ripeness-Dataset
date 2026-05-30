from __future__ import annotations

import argparse
import hashlib
import json
import math
import shutil
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageOps
from scipy import ndimage
from tqdm import tqdm


CLASS_INFO = [
    {
        "id": 1,
        "name": "green_immature",
        "display_name": "Green immature",
        "color_rgb": [35, 170, 70],
    },
    {
        "id": 2,
        "name": "pale_pink",
        "display_name": "Pale pink",
        "color_rgb": [255, 150, 190],
    },
    {
        "id": 3,
        "name": "pink_turns_purple",
        "display_name": "Pink turns purple",
        "color_rgb": [155, 70, 210],
    },
    {
        "id": 4,
        "name": "fully_ripe",
        "display_name": "Fully ripe",
        "color_rgb": [45, 105, 230],
    },
    {
        "id": 5,
        "name": "over_ripe",
        "display_name": "Over ripe",
        "color_rgb": [210, 95, 35],
    },
]

CLASS_NAMES = [item["name"] for item in CLASS_INFO]
CLASS_IDS = {item["name"]: item["id"] for item in CLASS_INFO}


@dataclass(frozen=True)
class ReleasePaths:
    root: Path
    dataset: Path
    images: Path
    binary_masks: Path
    semantic_masks: Path
    overall_masks: Path
    metadata: Path
    reports: Path
    figures: Path
    overlays: Path
    splits: Path


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def make_paths(output_root: Path) -> ReleasePaths:
    return ReleasePaths(
        root=output_root,
        dataset=output_root / "dataset",
        images=output_root / "dataset" / "images",
        binary_masks=output_root / "dataset" / "masks_binary",
        semantic_masks=output_root / "dataset" / "masks_semantic",
        overall_masks=output_root / "dataset" / "masks_overall",
        metadata=output_root / "metadata",
        reports=output_root / "reports",
        figures=output_root / "figures",
        overlays=output_root / "figures" / "overlays",
        splits=output_root / "splits",
    )


def reset_output(paths: ReleasePaths, force: bool) -> None:
    if paths.root.exists():
        if not force:
            raise FileExistsError(
                f"{paths.root} already exists. Pass --force to replace generated outputs."
            )
        shutil.rmtree(paths.root)
    for path in [
        paths.images,
        paths.semantic_masks,
        paths.overall_masks,
        paths.metadata,
        paths.reports,
        paths.figures,
        paths.overlays,
        paths.splits,
    ]:
        path.mkdir(parents=True, exist_ok=True)
    for class_name in CLASS_NAMES:
        (paths.binary_masks / class_name).mkdir(parents=True, exist_ok=True)


def read_count_sheet(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Count sheet is empty: {path}")
    headers = [value for value in rows[0] if value is not None]
    records = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        filename = str(row[0]).strip()
        if filename.lower() == "total":
            continue
        record = dict(zip(headers, row[: len(headers)]))
        for class_name in CLASS_NAMES:
            record[class_name] = int(record.get(class_name) or 0)
        record["Total"] = int(record.get("Total") or sum(record[class_name] for class_name in CLASS_NAMES))
        records.append(record)
    output = pd.DataFrame(records)
    if "Image Name" not in output.columns:
        raise ValueError("Expected count sheet column 'Image Name'")
    output = output.rename(columns={"Image Name": "filename", "Total": "total_objects"})
    return output[["filename", *CLASS_NAMES, "total_objects"]].sort_values("filename").reset_index(drop=True)


def list_input_images(data_root: Path) -> list[Path]:
    images_root = data_root / "images"
    if not images_root.exists():
        raise FileNotFoundError(f"Expected RGB images under {images_root}")
    images = sorted(images_root.glob("*.jpg"))
    if not images:
        raise FileNotFoundError(f"No JPG images found under {images_root}")
    return images


def validate_pairing(data_root: Path, image_paths: Iterable[Path]) -> None:
    missing = []
    for image_path in image_paths:
        for folder in ["Overall", *CLASS_NAMES]:
            mask_path = data_root / folder / image_path.name
            if not mask_path.exists():
                missing.append(str(mask_path))
    if missing:
        preview = "\n".join(missing[:20])
        raise FileNotFoundError(f"Missing image-mask pairs:\n{preview}")


def threshold_mask(mask: Image.Image, threshold: int) -> tuple[np.ndarray, int]:
    array = np.array(mask.convert("L"))
    ambiguous_pixels = int(((array > 0) & (array < 255)).sum())
    binary = array > threshold
    return binary, ambiguous_pixels


def save_binary_png(mask: np.ndarray, path: Path) -> None:
    image = Image.fromarray((mask.astype(np.uint8) * 255), mode="L")
    image.save(path, format="PNG", optimize=True)


def image_quality_metrics(image: Image.Image) -> dict[str, float]:
    small_rgb = ImageOps.contain(image, (512, 512), Image.Resampling.LANCZOS)
    gray = np.array(small_rgb.convert("L"), dtype=np.float32)
    laplace = ndimage.laplace(gray)
    hsv = np.array(small_rgb.convert("HSV"), dtype=np.float32)
    return {
        "brightness_mean": float(gray.mean()),
        "brightness_std": float(gray.std()),
        "sharpness_laplacian_var": float(laplace.var()),
        "saturation_mean": float(hsv[:, :, 1].mean()),
    }


def connected_component_count(mask: np.ndarray) -> int:
    _, count = ndimage.label(mask)
    return int(count)


def make_overlay(
    image: Image.Image,
    masks: dict[str, np.ndarray],
    output_path: Path,
    title: str,
    max_width: int = 1500,
) -> None:
    overlay = image.convert("RGBA")
    for item in CLASS_INFO:
        mask = masks[item["name"]]
        rgba = np.zeros((mask.shape[0], mask.shape[1], 4), dtype=np.uint8)
        rgba[mask, :3] = item["color_rgb"]
        rgba[mask, 3] = 105
        overlay = Image.alpha_composite(overlay, Image.fromarray(rgba, mode="RGBA"))

    scale = min(1.0, max_width / overlay.width)
    resized = overlay.resize(
        (int(overlay.width * scale), int(overlay.height * scale)),
        Image.Resampling.LANCZOS,
    )
    strip_height = 34 + 22 * len(CLASS_INFO)
    canvas = Image.new("RGB", (resized.width, resized.height + strip_height), "white")
    canvas.paste(resized.convert("RGB"), (0, 0))
    draw = ImageDraw.Draw(canvas)
    y = resized.height + 8
    draw.text((12, y), title, fill="black")
    y += 22
    for item in CLASS_INFO:
        color = tuple(item["color_rgb"])
        pixels = int(masks[item["name"]].sum())
        draw.rectangle([12, y + 3, 28, y + 17], fill=color)
        draw.text((36, y), f"{item['name']}: {pixels:,} positive pixels", fill="black")
        y += 22
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, quality=92)


def select_overlay_examples(manifest: pd.DataFrame, limit: int = 12) -> list[str]:
    selected: list[str] = []
    ranked_total = manifest.sort_values("total_objects", ascending=False)["filename"].head(3).tolist()
    selected.extend(ranked_total)
    for class_name in CLASS_NAMES:
        candidate = (
            manifest[manifest[f"{class_name}_objects"] > 0]
            .sort_values(f"{class_name}_objects", ascending=False)["filename"]
            .head(1)
            .tolist()
        )
        selected.extend(candidate)
    selected.extend(manifest.sample(n=min(limit, len(manifest)), random_state=42)["filename"].tolist())
    deduped = []
    for filename in selected:
        if filename not in deduped:
            deduped.append(filename)
        if len(deduped) >= limit:
            break
    return deduped


def make_contact_sheet(image_paths: list[Path], output_path: Path, columns: int = 4) -> None:
    if not image_paths:
        return
    thumb_w, thumb_h = 360, 480
    label_h = 30
    pad = 10
    rows = math.ceil(len(image_paths) / columns)
    tile_w = thumb_w + 2 * pad
    tile_h = thumb_h + label_h + 2 * pad
    canvas = Image.new("RGB", (columns * tile_w, rows * tile_h), "white")
    draw = ImageDraw.Draw(canvas)
    for index, path in enumerate(image_paths):
        x = (index % columns) * tile_w
        y = (index // columns) * tile_h
        image = Image.open(path).convert("RGB")
        image.thumbnail((thumb_w, thumb_h), Image.Resampling.LANCZOS)
        canvas.paste(image, (x + pad + (thumb_w - image.width) // 2, y + pad))
        draw.text((x + pad, y + pad + thumb_h + 4), path.stem, fill="black")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, quality=92)


def split_score(
    split_counts: dict[str, np.ndarray],
    split_sizes: dict[str, int],
    target_counts: dict[str, np.ndarray],
    target_sizes: dict[str, int],
) -> float:
    score = 0.0
    for split_name in target_sizes:
        denom = np.maximum(target_counts[split_name], 1.0)
        diff = (split_counts[split_name] - target_counts[split_name]) / denom
        score += float(np.square(diff).sum())
        size_diff = (split_sizes[split_name] - target_sizes[split_name]) / target_sizes[split_name]
        score += float(size_diff * size_diff)
    return score


def assign_balanced_splits(
    manifest: pd.DataFrame,
    seed: int,
    repeats: int,
) -> tuple[pd.Series, pd.DataFrame]:
    split_names = ["train", "val", "test"]
    target_sizes = {"train": 296, "val": 64, "test": 64}
    total_counts = manifest[[f"{class_name}_objects" for class_name in CLASS_NAMES]].sum().to_numpy(float)
    total_objects = manifest["total_objects"].to_numpy(float)
    target_counts = {name: total_counts * (target_sizes[name] / len(manifest)) for name in split_names}
    target_objects = {
        name: float(total_objects.sum() * (target_sizes[name] / len(manifest)))
        for name in split_names
    }
    vectors = manifest[[f"{class_name}_objects" for class_name in CLASS_NAMES]].to_numpy(float)

    def objective(assignment_codes: np.ndarray) -> float:
        score = 0.0
        for split_index, split_name in enumerate(split_names):
            mask = assignment_codes == split_index
            class_counts = vectors[mask].sum(axis=0)
            object_count = total_objects[mask].sum()
            class_error = (class_counts - target_counts[split_name]) / np.maximum(
                target_counts[split_name],
                1.0,
            )
            object_error = (object_count - target_objects[split_name]) / target_objects[split_name]
            score += float(np.square(class_error).sum() * 2.0)
            score += float(object_error * object_error * 5.0)
        return score

    best_assignment: np.ndarray | None = None
    best_score = float("inf")
    rng_master = np.random.default_rng(seed)
    base_assignment = np.array(
        [0] * target_sizes["train"] + [1] * target_sizes["val"] + [2] * target_sizes["test"],
        dtype=np.int8,
    )

    for _ in range(repeats):
        assignment = base_assignment.copy()
        rng_master.shuffle(assignment)
        score = objective(assignment)
        if score < best_score:
            best_score = score
            best_assignment = assignment.copy()

    if best_assignment is None:
        raise RuntimeError("Failed to create balanced split assignment")

    # Pairwise swaps refine the best random split while preserving exact split sizes.
    rng = np.random.default_rng(seed + 1)
    for _ in range(30):
        improved = False
        for left in rng.permutation(len(manifest)):
            for right in rng.permutation(len(manifest)):
                if best_assignment[left] == best_assignment[right]:
                    continue
                trial = best_assignment.copy()
                trial[left], trial[right] = trial[right], trial[left]
                score = objective(trial)
                if score + 1e-12 < best_score:
                    best_assignment = trial
                    best_score = score
                    improved = True
                    break
            if improved:
                break
        if not improved:
            break

    split_series = pd.Series(
        [split_names[int(code)] for code in best_assignment],
        index=manifest.index,
        name="split",
    )
    split_stats = (
        manifest.assign(split=split_series)
        .groupby("split")
        .agg(
            images=("filename", "count"),
            total_objects=("total_objects", "sum"),
            **{
                f"{class_name}_objects": (f"{class_name}_objects", "sum")
                for class_name in CLASS_NAMES
            },
        )
        .reindex(split_names)
        .reset_index()
    )
    return split_series, split_stats


def write_json(path: Path, data: object) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def plot_class_distribution(class_stats: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8, 4.8))
    ax.bar(class_stats["class_name"], class_stats["objects"], color="#4f7f58")
    ax.set_ylabel("Annotated berry instances")
    ax.set_xlabel("Ripeness class")
    ax.tick_params(axis="x", rotation=25)
    ax.set_title("Class distribution by object count")
    fig.tight_layout()
    fig.savefig(output_path, dpi=240)
    plt.close(fig)


def plot_split_distribution(split_stats: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    bottom = np.zeros(len(split_stats))
    x = np.arange(len(split_stats))
    for item in CLASS_INFO:
        values = split_stats[f"{item['name']}_objects"].to_numpy()
        ax.bar(x, values, bottom=bottom, label=item["name"])
        bottom += values
    ax.set_xticks(x)
    ax.set_xticklabels(split_stats["split"])
    ax.set_ylabel("Annotated berry instances")
    ax.set_title("Object distribution across recommended splits")
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(output_path, dpi=240)
    plt.close(fig)


def plot_quality_distribution(manifest: pd.DataFrame, output_path: Path) -> None:
    fig, axes = plt.subplots(1, 3, figsize=(11, 3.8))
    axes[0].hist(manifest["brightness_mean"], bins=24, color="#607d8b")
    axes[0].set_title("Brightness")
    axes[1].hist(manifest["sharpness_laplacian_var"], bins=24, color="#8d6e63")
    axes[1].set_title("Sharpness")
    axes[2].hist(manifest["saturation_mean"], bins=24, color="#789262")
    axes[2].set_title("Saturation")
    for ax in axes:
        ax.set_ylabel("Images")
    fig.tight_layout()
    fig.savefig(output_path, dpi=240)
    plt.close(fig)


def dataframe_to_markdown(df: pd.DataFrame, index: bool = False) -> str:
    return df.to_markdown(index=index)


def write_dataset_card(
    paths: ReleasePaths,
    manifest: pd.DataFrame,
    class_stats: pd.DataFrame,
    split_stats: pd.DataFrame,
) -> None:
    lines = [
        "# Silal Blueberry Ripeness Dataset Release",
        "",
        "This release package organizes the RGB blueberry images, five ripeness-stage binary masks, an overall berry mask, semantic label maps, image-level count metadata, recommended splits, and technical validation reports for a Scientific Data style manuscript.",
        "",
        "## Dataset Summary",
        "",
        f"- RGB greenhouse images: {len(manifest):,}",
        f"- Original acquisition device: {manifest['camera_make'].mode().iloc[0]} {manifest['camera_model'].mode().iloc[0]}",
        f"- Annotated berry instances from count sheet: {int(manifest['total_objects'].sum()):,}",
        "- Ripeness stages: green immature, pale pink, pink turns purple, fully ripe, over ripe.",
        "- Label format: per-class binary PNG masks plus one semantic PNG label map per image.",
        "- Semantic label values: 0 background, 1 green_immature, 2 pale_pink, 3 pink_turns_purple, 4 fully_ripe, 5 over_ripe, 255 overlap/ignore.",
        "",
        "## Data Structure",
        "",
        "```text",
        "scientific_data_release/",
        "  dataset/",
        "    images/                       # EXIF-oriented RGB JPG images",
        "    masks_binary/",
        "      green_immature/             # Binary PNG masks, 0 background, 255 positive",
        "      pale_pink/",
        "      pink_turns_purple/",
        "      fully_ripe/",
        "      over_ripe/",
        "    masks_overall/                # Binary PNG union/overall berry masks",
        "    masks_semantic/               # Single-channel semantic PNG masks",
        "  metadata/",
        "    class_map.json",
        "    dataset_manifest.csv",
        "    image_level_counts.csv",
        "    mask_overlap_audit.csv",
        "    split_summary.csv",
        "  splits/",
        "    train.txt",
        "    val.txt",
        "    test.txt",
        "  figures/",
        "    class_object_distribution.png",
        "    split_object_distribution.png",
        "    image_quality_distribution.png",
        "    overlays/",
        "  reports/",
        "    technical_validation.md",
        "    scientific_data_release_report.md",
        "    manuscript_data_descriptor_outline.md",
        "```",
        "",
        "## Class Distribution",
        "",
        dataframe_to_markdown(class_stats[["class_name", "objects", "object_percent", "images_with_objects"]]),
        "",
        "## Recommended Split Distribution",
        "",
        dataframe_to_markdown(split_stats),
        "",
        "## Usage Notes",
        "",
        "- Use `dataset/images/<filename>.jpg` as the model input.",
        "- Use `dataset/masks_binary/<class>/<filename>.png` for per-class semantic segmentation or multilabel segmentation.",
        "- Use `dataset/masks_semantic/<filename>.png` for single-head multiclass semantic segmentation.",
        "- Treat semantic value `255` as ignore if using the combined semantic masks.",
        "- Use the split text files for reproducible benchmark comparisons.",
        "",
    ]
    (paths.root / "DATASET_CARD.md").write_text("\n".join(lines), encoding="utf-8")


def write_technical_validation(
    paths: ReleasePaths,
    manifest: pd.DataFrame,
    class_stats: pd.DataFrame,
    split_stats: pd.DataFrame,
    overlap_audit: pd.DataFrame,
    union_mismatch: pd.Series,
    ambiguous_pixels: pd.Series,
) -> None:
    dimension_table = (
        manifest.groupby(["width", "height"]).size().reset_index(name="images").sort_values("images", ascending=False)
    )
    orientation_table = (
        manifest.groupby("exif_orientation").size().reset_index(name="images").sort_values("images", ascending=False)
    )
    empty_masks = []
    for class_name in CLASS_NAMES:
        empty_masks.append(
            {
                "class_name": class_name,
                "empty_masks": int((manifest[f"{class_name}_mask_pixels"] == 0).sum()),
                "non_empty_masks": int((manifest[f"{class_name}_mask_pixels"] > 0).sum()),
            }
        )
    empty_df = pd.DataFrame(empty_masks)
    quality_summary = manifest[
        ["brightness_mean", "brightness_std", "sharpness_laplacian_var", "saturation_mean"]
    ].describe().reset_index()

    lines = [
        "# Technical Validation",
        "",
        "## Completeness And File Integrity",
        "",
        f"- RGB image-mask groups checked: {len(manifest):,}",
        "- Every RGB image has one overall mask and five class-specific masks.",
        f"- Unique release image SHA-256 hashes: {manifest['release_sha256'].nunique():,}",
        f"- Unique filenames: {manifest['filename'].nunique():,}",
        "- No missing image-mask pairs were found during preprocessing.",
        "",
        "## Image Geometry And Orientation",
        "",
        "Images were EXIF-transposed before release so the RGB pixels align directly with the masks.",
        "",
        dataframe_to_markdown(dimension_table),
        "",
        dataframe_to_markdown(orientation_table),
        "",
        "## Annotation Counts",
        "",
        dataframe_to_markdown(class_stats[["class_name", "objects", "object_percent", "images_with_objects", "mask_pixels", "mask_area_percent"]]),
        "",
        "## Empty-Mask Audit",
        "",
        "Empty masks are expected for rare classes when a ripeness stage is absent from an image.",
        "",
        dataframe_to_markdown(empty_df),
        "",
        "## Mask Conversion Audit",
        "",
        "The source masks are JPEG files and therefore contain antialiasing/compression edge values. The release converts them to thresholded binary PNG masks using threshold > 127.",
        "",
        f"- Source mask ambiguous edge/compression pixels across class and overall masks: {int(ambiguous_pixels.sum()):,}",
        f"- Images with any class-overlap pixels after thresholding: {int((overlap_audit['overlap_pixels'] > 0).sum()):,}",
        f"- Maximum class-overlap pixels in one image: {int(overlap_audit['overlap_pixels'].max()):,}",
        f"- Mean class-overlap pixels per image: {overlap_audit['overlap_pixels'].mean():.2f}",
        f"- Images with any class-union vs overall-mask mismatch pixels: {int((union_mismatch > 0).sum()):,}",
        f"- Maximum union mismatch pixels in one image: {int(union_mismatch.max()):,}",
        f"- Mean union mismatch pixels per image: {union_mismatch.mean():.2f}",
        "",
        "Overlap pixels are encoded as value 255 in `masks_semantic` and should be treated as ignore pixels for semantic segmentation training. The original per-class binary masks remain available for multilabel segmentation.",
        "",
        "## Image Quality Screening",
        "",
        "Brightness, contrast, saturation, and Laplacian sharpness were measured on downsampled EXIF-oriented images to screen for gross acquisition problems.",
        "",
        dataframe_to_markdown(quality_summary),
        "",
        "## Recommended Benchmark Split",
        "",
        "A deterministic 296/64/64 train/validation/test split was generated to balance image counts and approximate object-count balance across ripeness stages.",
        "",
        dataframe_to_markdown(split_stats),
        "",
        "## Validation Figures",
        "",
        "- `figures/class_object_distribution.png`",
        "- `figures/split_object_distribution.png`",
        "- `figures/image_quality_distribution.png`",
        "- `figures/overlays/contact_sheet_overlays.jpg`",
        "",
    ]
    (paths.reports / "technical_validation.md").write_text("\n".join(lines), encoding="utf-8")


def write_manuscript_report(
    paths: ReleasePaths,
    manifest: pd.DataFrame,
    class_stats: pd.DataFrame,
    split_stats: pd.DataFrame,
) -> None:
    selling_points = [
        "High-resolution greenhouse imagery: 424 RGB images at phone-camera resolution, mostly 3000 x 4000 after EXIF orientation.",
        "Fine-grained ripeness taxonomy: five ripeness stages instead of only ripe/unripe or mature/immature.",
        "Dense annotation utility: 13,909 berry instances are counted at image level, with per-class pixel masks for segmentation.",
        "Multi-task potential: semantic segmentation, multilabel segmentation, berry counting, maturity distribution estimation, and harvest-readiness analysis.",
        "Release-readiness: lossless thresholded PNG masks, semantic maps, class map, split files, file hashes, QA tables, and visual overlays.",
    ]
    lines = [
        "# Scientific Data Release Report",
        "",
        "## Recommended Manuscript Framing",
        "",
        "Position the dataset as a high-resolution greenhouse blueberry ripeness segmentation and counting resource. Avoid framing it as only an image-level classification dataset; the stronger contribution is paired RGB imagery, five ripeness masks, and image-level berry counts.",
        "",
        "Suggested title:",
        "",
        "> A high-resolution greenhouse blueberry dataset with five-stage ripeness masks and berry count annotations",
        "",
        "## Core Quantitative Claims",
        "",
        f"- Images: {len(manifest):,}",
        f"- Annotated berries: {int(manifest['total_objects'].sum()):,}",
        f"- Classes: {len(CLASS_NAMES)} ripeness stages",
        f"- Image dimensions after orientation: {manifest.groupby(['width', 'height']).size().to_dict()}",
        f"- Acquisition device: {manifest['camera_make'].mode().iloc[0]} {manifest['camera_model'].mode().iloc[0]}",
        f"- Acquisition date range in EXIF: {manifest['datetime_original'].min()} to {manifest['datetime_original'].max()}",
        "",
        "## Class Distribution",
        "",
        dataframe_to_markdown(class_stats[["class_name", "objects", "object_percent", "images_with_objects"]]),
        "",
        "## Why This Is More Publishable Now",
        "",
    ]
    lines.extend([f"- {point}" for point in selling_points])
    lines.extend(
        [
            "",
            "## Recommended Scientific Data Sections",
            "",
            "### Background And Summary",
            "",
            "Explain the need for fine-grained blueberry ripeness estimation in controlled-environment agriculture, robotic harvesting, yield forecasting, and harvest scheduling. Emphasize that many existing blueberry datasets focus on detection or binary maturity, while this dataset provides five ripeness stages with pixel-level masks.",
            "",
            "### Methods",
            "",
            "Describe greenhouse acquisition, device, image resolution, sampling period, ripeness taxonomy, annotation workflow, mask export procedure, and preprocessing. Include the EXIF orientation correction and JPEG-mask-to-PNG conversion threshold.",
            "",
            "### Data Records",
            "",
            "Describe the release folder structure, file naming convention, class map, `dataset_manifest.csv`, `image_level_counts.csv`, and split files.",
            "",
            "### Technical Validation",
            "",
            "Use `reports/technical_validation.md` as the backbone. Include completeness checks, image-mask alignment, class distributions, mask conversion QA, overlap handling, quality metrics, and representative overlays.",
            "",
            "### Usage Notes",
            "",
            "Document semantic label IDs, ignore value 255, and recommended tasks: semantic segmentation, ripeness counting, maturity distribution estimation, and benchmark split evaluation.",
            "",
            "## Additional Work That Would Increase Acceptance Probability",
            "",
            "- Add a small independent annotation audit: randomly sample 40 to 60 berries and report expert agreement or correction rate.",
            "- Add one baseline semantic segmentation experiment, preferably DeepLabV3+, U-Net, SegFormer, or Mask2Former, reporting mIoU/Dice per ripeness class.",
            "- Add a counting validation baseline: predicted class counts from masks or a trained model versus spreadsheet counts, using MAE/RMSE by class.",
            "- Replace or supplement JPEG-derived masks with original annotation exports if available, such as polygon JSON, LabelMe, CVAT, COCO, or uncompressed PNG exports.",
            "- Publish the release with a DOI and explicit license, plus a datasheet covering consent, privacy, and intended use.",
            "",
            "## Recommended Split Summary",
            "",
            dataframe_to_markdown(split_stats),
            "",
        ]
    )
    report_text = "\n".join(lines)
    (paths.reports / "scientific_data_release_report.md").write_text(report_text, encoding="utf-8")

    outline = [
        "# Manuscript Data Descriptor Outline",
        "",
        "## Title",
        "",
        "A high-resolution greenhouse blueberry dataset with five-stage ripeness masks and berry count annotations",
        "",
        "## Abstract Skeleton",
        "",
        "We present a high-resolution greenhouse blueberry image dataset containing 424 RGB images and five ripeness-stage mask layers. The dataset includes image-level counts for 13,909 annotated berries across green immature, pale pink, pink-turns-purple, fully ripe, and over-ripe stages. We provide lossless binary masks, semantic label maps, metadata, recommended train/validation/test splits, and technical validation reports to support ripeness segmentation, counting, and harvest-readiness analysis.",
        "",
        "## Figure Plan",
        "",
        "1. Acquisition and annotation overview.",
        "2. Dataset structure and class taxonomy.",
        "3. Representative RGB images with ripeness mask overlays.",
        "4. Class/object distribution and split distribution.",
        "5. Technical validation summary: image quality, mask integrity, and overlap/mismatch audit.",
        "",
        "## Table Plan",
        "",
        "1. Dataset record summary.",
        "2. Class distribution by object count and image presence.",
        "3. File structure and data records.",
        "4. Technical validation checks.",
        "5. Recommended split distribution.",
        "",
    ]
    (paths.reports / "manuscript_data_descriptor_outline.md").write_text("\n".join(outline), encoding="utf-8")


def copy_report_to_docs(paths: ReleasePaths, docs_output: Path | None) -> None:
    if docs_output is None:
        return
    docs_output.parent.mkdir(parents=True, exist_ok=True)
    source = paths.reports / "scientific_data_release_report.md"
    docs_output.write_text(source.read_text(encoding="utf-8"), encoding="utf-8")


def process_release(data_root: Path, paths: ReleasePaths, threshold: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    count_df = read_count_sheet(data_root / "Image Wise Classname Count.xlsx")
    image_paths = list_input_images(data_root)
    validate_pairing(data_root, image_paths)
    count_lookup = count_df.set_index("filename").to_dict(orient="index")

    rows = []
    overlap_rows = []
    ambiguous_rows = []
    overlay_cache: dict[str, tuple[Image.Image, dict[str, np.ndarray]]] = {}

    for image_path in tqdm(image_paths, desc="preprocess release"):
        filename = image_path.name
        if filename not in count_lookup:
            raise KeyError(f"{filename} is missing from the count sheet")

        with Image.open(image_path) as raw_image:
            exif = raw_image.getexif()
            raw_width, raw_height = raw_image.size
            exif_orientation = int(exif.get(274, 1) or 1)
            camera_make = str(exif.get(271, "") or "")
            camera_model = str(exif.get(272, "") or "")
            datetime_original = str(exif.get(36867, "") or exif.get(306, "") or "")
            image = ImageOps.exif_transpose(raw_image).convert("RGB")

        release_image_path = paths.images / filename
        image.save(release_image_path, format="JPEG", quality=95, optimize=True)
        width, height = image.size
        quality = image_quality_metrics(image)

        masks: dict[str, np.ndarray] = {}
        mask_pixels = {}
        component_counts = {}
        ambiguous_total = 0

        for class_name in CLASS_NAMES:
            source_mask_path = data_root / class_name / filename
            with Image.open(source_mask_path) as source_mask:
                mask, ambiguous_pixels = threshold_mask(source_mask, threshold)
            if mask.shape != (height, width):
                raise ValueError(
                    f"Mask shape does not match EXIF-oriented image for {filename}: "
                    f"{class_name} mask {mask.shape}, image {(height, width)}"
                )
            masks[class_name] = mask
            save_binary_png(mask, paths.binary_masks / class_name / f"{image_path.stem}.png")
            mask_pixels[class_name] = int(mask.sum())
            component_counts[class_name] = connected_component_count(mask)
            ambiguous_total += ambiguous_pixels

        with Image.open(data_root / "Overall" / filename) as source_overall:
            overall_mask, overall_ambiguous = threshold_mask(source_overall, threshold)
        save_binary_png(overall_mask, paths.overall_masks / f"{image_path.stem}.png")
        ambiguous_total += overall_ambiguous

        stack = np.stack([masks[class_name] for class_name in CLASS_NAMES], axis=0)
        class_sum = stack.sum(axis=0)
        overlap_mask = class_sum > 1
        union_mask = class_sum > 0
        union_mismatch_pixels = int(np.logical_xor(union_mask, overall_mask).sum())
        overlap_pixels = int(overlap_mask.sum())

        semantic = np.zeros((height, width), dtype=np.uint8)
        for class_name in CLASS_NAMES:
            semantic[masks[class_name] & ~overlap_mask] = CLASS_IDS[class_name]
        semantic[overlap_mask] = 255
        Image.fromarray(semantic, mode="L").save(paths.semantic_masks / f"{image_path.stem}.png", format="PNG", optimize=True)

        counts = count_lookup[filename]
        row = {
            "image_id": image_path.stem,
            "filename": filename,
            "image_path": f"dataset/images/{filename}",
            "semantic_mask_path": f"dataset/masks_semantic/{image_path.stem}.png",
            "overall_mask_path": f"dataset/masks_overall/{image_path.stem}.png",
            "raw_width": raw_width,
            "raw_height": raw_height,
            "width": width,
            "height": height,
            "exif_orientation": exif_orientation,
            "camera_make": camera_make,
            "camera_model": camera_model,
            "datetime_original": datetime_original,
            "raw_bytes": image_path.stat().st_size,
            "release_bytes": release_image_path.stat().st_size,
            "raw_sha256": sha256_file(image_path),
            "release_sha256": sha256_file(release_image_path),
            "total_objects": int(counts["total_objects"]),
            "overall_mask_pixels": int(overall_mask.sum()),
            "union_mismatch_pixels": union_mismatch_pixels,
            "overlap_pixels": overlap_pixels,
            "source_mask_ambiguous_pixels": ambiguous_total,
            **quality,
        }
        for class_name in CLASS_NAMES:
            row[f"{class_name}_objects"] = int(counts[class_name])
            row[f"{class_name}_mask_path"] = f"dataset/masks_binary/{class_name}/{image_path.stem}.png"
            row[f"{class_name}_mask_pixels"] = mask_pixels[class_name]
            row[f"{class_name}_component_count"] = component_counts[class_name]
            row[f"{class_name}_coverage_percent"] = mask_pixels[class_name] / (width * height) * 100.0
        rows.append(row)

        overlap_rows.append(
            {
                "filename": filename,
                "overlap_pixels": overlap_pixels,
                "overlap_percent": overlap_pixels / (width * height) * 100.0,
                "union_mismatch_pixels": union_mismatch_pixels,
                "union_mismatch_percent": union_mismatch_pixels / (width * height) * 100.0,
            }
        )
        ambiguous_rows.append({"filename": filename, "source_mask_ambiguous_pixels": ambiguous_total})

        if len(overlay_cache) < 20:
            overlay_cache[filename] = (image.copy(), {name: value.copy() for name, value in masks.items()})

    manifest = pd.DataFrame(rows).sort_values("filename").reset_index(drop=True)
    split_series, split_stats = assign_balanced_splits(manifest, seed=42, repeats=20000)
    manifest["split"] = split_series

    # Reorder frequently used columns to the front.
    front_cols = [
        "image_id",
        "filename",
        "split",
        "image_path",
        "semantic_mask_path",
        "overall_mask_path",
        "width",
        "height",
        "total_objects",
    ]
    remaining = [column for column in manifest.columns if column not in front_cols]
    manifest = manifest[front_cols + remaining]

    count_df.to_csv(paths.metadata / "image_level_counts.csv", index=False)
    pd.DataFrame(overlap_rows).to_csv(paths.metadata / "mask_overlap_audit.csv", index=False)
    pd.DataFrame(ambiguous_rows).to_csv(paths.metadata / "source_mask_ambiguous_pixels.csv", index=False)
    manifest.to_csv(paths.metadata / "dataset_manifest.csv", index=False)
    split_stats.to_csv(paths.metadata / "split_summary.csv", index=False)

    for split_name in ["train", "val", "test"]:
        filenames = manifest[manifest["split"] == split_name]["filename"].sort_values()
        (paths.splits / f"{split_name}.txt").write_text("\n".join(filenames) + "\n", encoding="utf-8")

    class_rows = []
    total_objects = int(manifest["total_objects"].sum())
    total_pixels = int((manifest["width"] * manifest["height"]).sum())
    for class_name in CLASS_NAMES:
        objects = int(manifest[f"{class_name}_objects"].sum())
        mask_pixels = int(manifest[f"{class_name}_mask_pixels"].sum())
        class_rows.append(
            {
                "class_id": CLASS_IDS[class_name],
                "class_name": class_name,
                "objects": objects,
                "object_percent": objects / total_objects * 100.0,
                "images_with_objects": int((manifest[f"{class_name}_objects"] > 0).sum()),
                "mask_pixels": mask_pixels,
                "mask_area_percent": mask_pixels / total_pixels * 100.0,
                "connected_components": int(manifest[f"{class_name}_component_count"].sum()),
            }
        )
    class_stats = pd.DataFrame(class_rows)
    class_stats.to_csv(paths.metadata / "class_distribution.csv", index=False)

    write_json(
        paths.metadata / "class_map.json",
        {
            "background": {"id": 0, "name": "background"},
            "classes": CLASS_INFO,
            "semantic_ignore_value": 255,
            "binary_mask_positive_value": 255,
            "binary_mask_background_value": 0,
            "source_mask_threshold": threshold,
        },
    )

    plot_class_distribution(class_stats, paths.figures / "class_object_distribution.png")
    plot_split_distribution(split_stats, paths.figures / "split_object_distribution.png")
    plot_quality_distribution(manifest, paths.figures / "image_quality_distribution.png")

    overlay_names = select_overlay_examples(manifest, limit=12)
    overlay_paths = []
    for filename in overlay_names:
        if filename in overlay_cache:
            image, masks = overlay_cache[filename]
        else:
            image = Image.open(paths.images / filename).convert("RGB")
            masks = {
                class_name: np.array(Image.open(paths.binary_masks / class_name / f"{Path(filename).stem}.png").convert("L")) > 127
                for class_name in CLASS_NAMES
            }
        output_path = paths.overlays / f"{Path(filename).stem}_overlay.jpg"
        make_overlay(image, masks, output_path, title=f"{filename} with ripeness masks")
        overlay_paths.append(output_path)
    make_contact_sheet(overlay_paths, paths.overlays / "contact_sheet_overlays.jpg")

    overlap_audit = pd.read_csv(paths.metadata / "mask_overlap_audit.csv")
    ambiguous_series = manifest["source_mask_ambiguous_pixels"]
    union_mismatch = manifest["union_mismatch_pixels"]

    write_dataset_card(paths, manifest, class_stats, split_stats)
    write_technical_validation(paths, manifest, class_stats, split_stats, overlap_audit, union_mismatch, ambiguous_series)
    write_manuscript_report(paths, manifest, class_stats, split_stats)

    return manifest, class_stats, split_stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare the Silal blueberry dataset for Scientific Data release.")
    parser.add_argument("--data-root", default="data", help="Input data root containing images and mask folders.")
    parser.add_argument("--output-root", default="outputs/scientific_data_release", help="Output release root.")
    parser.add_argument("--threshold", type=int, default=127, help="Mask threshold for JPEG-to-binary PNG conversion.")
    parser.add_argument("--force", action="store_true", help="Replace the existing generated release folder.")
    parser.add_argument("--docs-output", default="docs/scientific_data_release_report.md", help="Optional copy of the release report.")
    args = parser.parse_args()

    data_root = Path(args.data_root).resolve()
    output_root = Path(args.output_root).resolve()
    paths = make_paths(output_root)
    reset_output(paths, force=args.force)

    manifest, class_stats, split_stats = process_release(data_root, paths, args.threshold)
    copy_report_to_docs(paths, Path(args.docs_output) if args.docs_output else None)

    print(f"Prepared Scientific Data release at: {paths.root}")
    print(f"Images: {len(manifest):,}")
    print(f"Annotated berry instances: {int(manifest['total_objects'].sum()):,}")
    print(f"Class distribution written to: {paths.metadata / 'class_distribution.csv'}")
    print(f"Dataset manifest written to: {paths.metadata / 'dataset_manifest.csv'}")
    print(f"Technical validation written to: {paths.reports / 'technical_validation.md'}")


if __name__ == "__main__":
    main()
